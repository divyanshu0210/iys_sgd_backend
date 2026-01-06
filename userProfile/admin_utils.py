import datetime
import logging
import openpyxl
from django.http import HttpResponse
from django.utils.html import strip_tags
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils.timezone import localtime, is_aware


logger = logging.getLogger(__name__)


def export_as_excel(modeladmin, request, queryset):
    """
    Admin action:
    - NO selection required
    - Exports ALL rows
    - Respects filters, search, ordering
    """

    try:
        if hasattr(modeladmin, "get_changelist_instance"):
            changelist = modeladmin.get_changelist_instance(request)
            queryset = changelist.get_queryset(request)
        else:
            queryset = modeladmin.get_queryset(request)

    except Exception:
        logger.exception("Failed to rebuild queryset for Excel export")
        return None

    logger.info(
        "Excel export (ALL rows) | model=%s | rows=%s",
        modeladmin.model._meta.model_name,
        queryset.count(),
    )

    return _export_queryset_as_excel(modeladmin, request, queryset)


def _export_queryset_as_excel(modeladmin, request, queryset):
    meta = modeladmin.model._meta

    list_display = [
        f for f in modeladmin.get_list_display(request)
        if f != "action_checkbox"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta.verbose_name_plural.title()

    # Header
    headers = []
    for field in list_display:
        if hasattr(modeladmin, field):
            attr = getattr(modeladmin, field)
            headers.append(
                getattr(attr, "short_description", field.replace("_", " ").title())
            )
        else:
            headers.append(field.replace("_", " ").title())

    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # nice blue
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    # Rows
    for obj in queryset.iterator(chunk_size=2000):
        row = []

        for field in list_display:
            try:
                if hasattr(modeladmin, field):
                    attr = getattr(modeladmin, field)
                    try:
                        value = attr(obj)
                    except TypeError:
                        value = attr(request, obj)
                elif hasattr(obj, field):
                    attr = getattr(obj, field)
                    value = attr() if callable(attr) else attr
                else:
                    value = ""
            except Exception:
                logger.exception(
                    "Excel cell error | model=%s | id=%s | field=%s",
                    meta.model_name,
                    obj.pk,
                    field,
                )
                value = ""

            # Excel-friendly formatting
            # ----------------------
            if isinstance(value, datetime.datetime):
                # Convert to local naive datetime for Excel
                if is_aware(value):
                    value = localtime(value)
                value = value.replace(tzinfo=None)
                cell_value = value
            elif isinstance(value, datetime.date):
                cell_value = value
            else:
                cell_value = strip_tags(str(value)) if value is not None else ""


            # row.append(strip_tags(str(value)) if value is not None else "")
            
            row.append(cell_value)

        ws.append(row)
    
     # ----------------------
    # Auto-adjust column widths & wrap text
    # ----------------------
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # get the column name (A, B, C...)
        for cell in col:
            if cell.value:
                # Convert dates to strings for width calculation
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    length = len(cell.value.strftime("%d %b %Y %H:%M"))
                else:
                    length = len(str(cell.value))
                max_length = max(max_length, length)
        adjusted_width = min(max_length + 2, 50)  # max width 50
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Apply Excel date format if it's a date/datetime
            if isinstance(cell.value, datetime.datetime):
                cell.number_format = "DD MMM YYYY HH:MM"
            elif isinstance(cell.value, datetime.date):
                cell.number_format = "DD MMM YYYY"
        ws.column_dimensions[column].width = adjusted_width


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{meta.verbose_name_plural}.xlsx"'
    )
    wb.save(response)
    return response
